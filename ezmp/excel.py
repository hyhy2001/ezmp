from typing import Callable, Optional, Any, List, Iterator
from .dataframe import map_df, _check_pandas, DataFrame, pd  # type: ignore


def _create_ref_getter(filepath: str, cache: Any):
    """
    Creates a callable that fetches references natively for a specific subprocess.
    """
    import openpyxl  # type: ignore

    def ref_getter(ref_str: str):
        hit = cache.get(ref_str) if cache else None
        if hit is not None:
            return hit

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
        sheet_name = wb.active.title
        cell_ref = ref_str
        if "!" in ref_str:
            sheet_name, cell_ref = ref_str.split("!")

        try:
            sheet = wb[sheet_name]
            if ":" in cell_ref:
                cells = sheet[cell_ref]
                if (
                    isinstance(cells, tuple)
                    and len(cells) > 0
                    and isinstance(cells[0], tuple)
                ):
                    val = [[c.value for c in r] for r in cells]
                elif isinstance(cells, tuple):
                    val = [c.value for c in cells]
                else:
                    val = cells.value
            else:
                val = sheet[cell_ref].value
        except Exception:
            val = None
        finally:
            wb.close()

        if cache:
            cache.set(ref_str, val)
        return val

    return ref_getter


def _make_eval_wrapper(user_func: Callable, filepath: str, cache: Any):
    """
    Wraps the user's DataFrame `.apply` target function, dynamically evaluating formulas
    on the row PRIOR to passing the row to the user function.
    """

    def wrapper(row):
        from ezmp.formula import evaluate_formula_string
        from ezmp.formula.errors import ExcelError

        # Instantiate ref_getter inside the subprocess
        ref_getter = _create_ref_getter(filepath, cache)

        for col in row.index:
            val = row[col]
            if isinstance(val, str) and str(val).startswith("="):
                res = evaluate_formula_string(str(val), ref_getter)
                if isinstance(res, ExcelError):
                    res = res.code
                row[col] = res

        return user_func(row)

    return wrapper


def map_excel(
    target_func: Callable,
    file_path: str,
    output_path: Optional[str] = None,
    use_threads: bool = False,
    max_workers: Optional[int] = None,
    desc: str = "Processing Excel rows",
) -> DataFrame:
    """
    Reads an Excel file, processes its rows concurrently, and optionally saves the result.
    """
    _check_pandas()

    df = pd.read_excel(file_path)
    result_df = map_df(
        target_func=target_func,
        df=df,
        use_threads=use_threads,
        max_workers=max_workers,
        desc=desc,
    )

    if output_path is not None:
        result_df.to_excel(output_path, index=False)

    return result_df


def _process_single_excel(file_path, target_func, read_kwargs):
    df = pd.read_excel(file_path, **read_kwargs)
    return target_func(df)


def map_excel_chunks(
    target_func: Callable,
    file_path: str,
    chunksize: int = 1000,
    use_threads: bool = False,
    max_workers: Optional[int] = None,
    desc: str = "Processing Excel chunks",
    evaluate_formulas: bool = False,
) -> Iterator[DataFrame]:
    """
    Reads an Excel file lazily in chunks, to prevent Out-Of-Memory (OOM) crashes
    on massive data matrices (e.g., millions of cells).
    If evaluate_formulas is True, ezmp dynamically parses Excel formulas (like VLOOKUP)
    via a native AST engine over parallel processes.

    Returns a Generator yielding processed DataFrame chunks.
    Dependencies: openpyxl
    """
    _check_pandas()
    import openpyxl  # type: ignore

    global_cache = None
    if evaluate_formulas:
        data_only = False
        from ezmp.cache import GlobalCache

        global_cache = GlobalCache(enabled=True)
    else:
        data_only = True

    actual_target = (
        _make_eval_wrapper(target_func, file_path, global_cache)
        if evaluate_formulas
        else target_func
    )

    def chunk_generator():
        # Use read_only=True for streaming, drastically reducing memory
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=data_only)
        ws = wb.active

        # Get headers
        rows_iter = ws.iter_rows(values_only=True)  # type: ignore
        headers = next(rows_iter)

        chunk_data = []
        chunk_index = 1

        for row in rows_iter:
            chunk_data.append(row)
            if len(chunk_data) >= chunksize:
                # Convert this chunk to a DataFrame
                df_chunk = pd.DataFrame(chunk_data, columns=headers)

                # Apply the function concurrently to the rows in this chunk
                processed_df = map_df(
                    target_func=actual_target,
                    df=df_chunk,
                    use_threads=use_threads,
                    max_workers=max_workers,
                    desc=f"{desc} (chunk {chunk_index})",
                )
                yield processed_df

                chunk_data = []
                chunk_index = chunk_index + 1  # type: ignore

        # Process the final remaining chunk if any
        if chunk_data:
            df_chunk = pd.DataFrame(chunk_data, columns=headers)
            yield map_df(
                target_func=actual_target,
                df=df_chunk,
                use_threads=use_threads,
                max_workers=max_workers,
                desc=f"{desc} (chunk {chunk_index})",
            )

        wb.close()

    return chunk_generator()


def map_excel_files(
    target_func: Callable[[DataFrame], Any],
    directory: str,
    recursive: bool = False,
    use_threads: bool = False,
    max_workers: Optional[int] = None,
    desc: str = "Scraping Excel files",
    **read_excel_kwargs,
) -> List[Any]:
    """
    Finds all Excel (.xlsx, .xls) files in a directory and applies `target_func`
    to each loaded DataFrame concurrently.

    Args:
        target_func: Function to apply to each loaded pd.DataFrame.
        directory: The directory containing Excel files.
        recursive: Whether to search subdirectories.
        use_threads: If True, uses threads; otherwise processes.
        max_workers: Max workers for concurrent execution.
        desc: Progress bar description.
        **read_excel_kwargs: Additional arguments passed to `pd.read_excel`.

    Returns:
        A list of results from `target_func`.
    """
    from . import files  # type: ignore
    import functools

    wrapper = functools.partial(
        _process_single_excel, target_func=target_func, read_kwargs=read_excel_kwargs
    )

    # Use files.map_dir to prevent glob logic duplication
    return files.map_dir(
        target_func=wrapper,
        dir_path=directory,
        pattern="*.xls*" if not recursive else "**/*.xls*",
        recursive=recursive,
        use_threads=use_threads,
        max_workers=max_workers,
        desc=desc,
    )
